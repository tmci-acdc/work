from openpyxl import load_workbook
print('#####################################')
print('#                                   #')
print('#                                   #')
print('#                                   #')
print('#####################################')
name = input("File name :")
fileName = "./" + name + '.xlsx'
fileNameNew = "./new_" + name + '.xlsx'

wb = load_workbook(fileName)

sheet = wb.active
# sheet = wb.worksheets('Арендаторы')

rowVal = [[9, 29], [43, 467]]

for k in range(len(rowVal)):
    for i in range(rowVal[k][0], rowVal[k][1]):
        oldVal = sheet.cell(row=i, column=5).value
        sheet.cell(row=i, column=6).value = oldVal
        sheet.cell(row=i, column=5).value = ''

sheet.column_dimensions.group('H', 'AD', hidden=True)
sheet.column_dimensions.group('A', hidden=True)
sheet.column_dimensions.group('D', hidden=True)
sheet.row_dimensions.group(467, 623, hidden=True)

print('Ok!')

wb.save(fileNameNew)
wb.close()
